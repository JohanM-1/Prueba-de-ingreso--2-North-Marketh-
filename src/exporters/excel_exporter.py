"""
Exportador de datos a Excel.
"""

import pandas as pd
from pathlib import Path
from typing import Dict, List, Any
from datetime import datetime
import os

from ..config.settings import OUTPUT_SETTINGS, DATA_PATHS
from ..utils.helpers import format_timestamp, sanitize_filename


class ExcelExporter:
    """
    Exporta datos de seguidores de Instagram a formato Excel.
    """
    
    def __init__(self, output_dir: str = None):
        """
        Inicializa el exportador.
        
        Args:
            output_dir: Directorio de salida (opcional)
        """
        self.output_dir = Path(output_dir) if output_dir else Path(DATA_PATHS['output'])
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_excel(
        self, 
        data: Dict[str, List[Dict[str, Any]]], 
        filename: str = None
    ) -> str:
        """
        Exporta datos a archivo Excel con múltiples hojas.
        
        Args:
            data: Datos por cuenta {account: [follower_data]}
            filename: Nombre del archivo (opcional)
            
        Returns:
            Ruta del archivo generado
        """
        if filename is None:
            timestamp = format_timestamp()
            filename = f"instagram_followers_{timestamp}.xlsx"
        
        filename = sanitize_filename(filename)
        output_path = self.output_dir / filename
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Crear hoja por cada cuenta
                for account, followers_data in data.items():
                    if not followers_data:
                        continue
                    
                    sheet_name = self._create_sheet_name(account)
                    
                    # Convertir a DataFrame
                    df = self._create_dataframe(followers_data)
                    
                    # Escribir a Excel
                    df.to_excel(
                        writer, 
                        sheet_name=sheet_name,
                        index=False,
                        freeze_panes=(1, 0)  # Congelar primera fila
                    )
                    
                    # Formatear hoja
                    self._format_worksheet(writer.book[sheet_name], df)
                
                # Crear hoja de resumen
                summary_df = self._create_summary_dataframe(data)
                summary_df.to_excel(writer, sheet_name='Resumen', index=False)
                self._format_summary_worksheet(writer.book['Resumen'], summary_df)
                
                # Crear hoja de metadatos
                if OUTPUT_SETTINGS['include_metadata']:
                    metadata_df = self._create_metadata_dataframe(data)
                    metadata_df.to_excel(writer, sheet_name='Metadatos', index=False)
            
            return str(output_path)
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise
    
    def _create_sheet_name(self, account: str) -> str:
        """
        Crea nombre de hoja válido para Excel.
        
        Args:
            account: Nombre de la cuenta
            
        Returns:
            Nombre de hoja sanitizado
        """
        sheet_name = f"{OUTPUT_SETTINGS['sheet_prefix']}{account}"
        
        # Excel limita nombres de hoja a 31 caracteres
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:28] + "..."
        
        # Eliminar caracteres problemáticos
        invalid_chars = ['\\', '/', '*', '[', ']', ':', '?']
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, '_')
            
        return sheet_name
    
    def _create_dataframe(self, followers_data: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Convierte datos de seguidores a DataFrame.
        
        Args:
            followers_data: Lista de datos de seguidores
            
        Returns:
            DataFrame con datos formateados
        """
        if not followers_data:
            return pd.DataFrame()
        
        # Normalizar datos para DataFrame
        normalized_data = []
        for follower in followers_data:
            normalized_follower = {}
            
            for key, value in follower.items():
                if isinstance(value, list):
                    # Convertir listas a strings separados por "; "
                    normalized_follower[key] = "; ".join(str(v) for v in value) if value else ""
                elif value is None:
                    normalized_follower[key] = ""
                else:
                    normalized_follower[key] = value
            
            normalized_data.append(normalized_follower)
        
        df = pd.DataFrame(normalized_data)
        
        # Reordenar columnas según importancia
        column_order = [
            'username', 'full_name', 'bio', 
            'posts_count', 'follower_count', 'following_count',
            'extraction_timestamp', 'source_account'
        ]
        
        # Mantener solo columnas que existen
        existing_columns = [col for col in column_order if col in df.columns]
        remaining_columns = [col for col in df.columns if col not in existing_columns]
        
        final_order = existing_columns + remaining_columns
        df = df[final_order]
        
        return df
    
    def _format_worksheet(self, worksheet, df: pd.DataFrame) -> None:
        """
        Aplica formato a hoja de Excel.
        
        Args:
            worksheet: Hoja de Excel
            df: DataFrame con datos
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Formatear encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Ajustar ancho de columnas
        for col_num, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            
            # Calcular ancho óptimo
            max_length = max(
                len(str(column)),  # Longitud del encabezado
                df[column].astype(str).str.len().max() if len(df) > 0 else 0
            )
            
            # Limitar ancho máximo
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Aplicar filtros automáticos
        if len(df) > 0:
            worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"
    
    def _create_summary_dataframe(self, data: Dict[str, List[Dict[str, Any]]]) -> pd.DataFrame:
        """
        Crea DataFrame de resumen.
        
        Args:
            data: Datos completos por cuenta
            
        Returns:
            DataFrame con resumen por cuenta
        """
        summary_data = []
        
        for account, followers_data in data.items():
            if not followers_data:
                continue
                
            # Calcular estadísticas
            total_followers = len(followers_data)
            with_phone = sum(1 for f in followers_data if f.get('phone_numbers'))
            verified_count = sum(1 for f in followers_data if f.get('is_verified'))
            private_count = sum(1 for f in followers_data if f.get('is_private'))
            
            summary_data.append({
                'Cuenta': f"@{account}",
                'Total_Seguidores_Extraídos': total_followers,
                'Con_Teléfono': with_phone,
                'Verificados': verified_count,
                'Privados': private_count,
                'Porcentaje_Teléfono': f"{(with_phone/total_followers)*100:.1f}%" if total_followers > 0 else "0%"
            })
        
        return pd.DataFrame(summary_data)
    
    def _format_summary_worksheet(self, worksheet, df: pd.DataFrame) -> None:
        """
        Aplica formato especial a hoja de resumen.
        
        Args:
            worksheet: Hoja de Excel
            df: DataFrame de resumen
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Formatear encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Ajustar anchos
        for col_num, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = 20
    
    def _create_metadata_dataframe(self, data: Dict[str, List[Dict[str, Any]]]) -> pd.DataFrame:
        """
        Crea DataFrame con metadatos de extracción.
        
        Args:
            data: Datos completos
            
        Returns:
            DataFrame con metadatos
        """
        metadata = [
            {'Campo': 'Fecha_Extracción', 'Valor': datetime.now().strftime(OUTPUT_SETTINGS['datetime_format'])},
            {'Campo': 'Total_Cuentas_Procesadas', 'Valor': len(data)},
            {'Campo': 'Total_Seguidores_Extraídos', 'Valor': sum(len(followers) for followers in data.values())},
            {'Campo': 'Formato_Fecha', 'Valor': OUTPUT_SETTINGS['date_format']},
            {'Campo': 'Versión_Extractor', 'Valor': '1.0.0'},
            {'Campo': 'Cumplimiento_GDPR', 'Valor': 'Solo datos públicos'},
            {'Campo': 'Campos_Obligatorios', 'Valor': 'username, full_name, is_private, extraction_timestamp, source_account'},
        ]
        
        return pd.DataFrame(metadata)
    
    def export_to_csv(
        self, 
        data: Dict[str, List[Dict[str, Any]]], 
        output_dir: str = None
    ) -> List[str]:
        """
        Exporta datos a archivos CSV separados por cuenta.
        
        Args:
            data: Datos por cuenta
            output_dir: Directorio de salida
            
        Returns:
            Lista de archivos CSV generados
        """
        if output_dir is None:
            output_dir = self.output_dir
        else:
            output_dir = Path(output_dir)
            output_dir.mkdir(parents=True, exist_ok=True)
        
        csv_files = []
        timestamp = format_timestamp()
        
        for account, followers_data in data.items():
            if not followers_data:
                continue
            
            filename = f"seguidores_{account}_{timestamp}.csv"
            csv_path = output_dir / filename
            
            df = self._create_dataframe(followers_data)
            df.to_csv(csv_path, index=False, encoding='utf-8-sig')
            
            csv_files.append(str(csv_path))
        
        return csv_files 